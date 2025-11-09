#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件分析脚本
用于分析保批单业务报表的结构和内容
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np
from datetime import datetime
import os

def analyze_excel_file(file_path):
    """
    分析Excel文件的详细结构和内容
    """
    print(f"正在分析文件: {file_path}")
    print("=" * 80)
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"错误: 文件 {file_path} 不存在")
        return
    
    try:
        # 使用openpyxl获取工作表信息
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        
        print(f"文件大小: {os.path.getsize(file_path) / 1024:.2f} KB")
        print(f"工作表数量: {len(workbook.sheetnames)}")
        print(f"工作表名称: {workbook.sheetnames}")
        print("\n" + "=" * 80 + "\n")
        
        # 分析每个工作表
        for sheet_name in workbook.sheetnames:
            print(f"工作表: {sheet_name}")
            print("-" * 60)
            
            # 使用pandas读取工作表数据
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # 基本统计信息
                print(f"数据行数: {len(df)}")
                print(f"数据列数: {len(df.columns)}")
                
                # 列信息
                print(f"\n列信息:")
                for i, col in enumerate(df.columns, 1):
                    print(f"  {i:2d}. {col}")
                
                # 数据类型分析
                print(f"\n数据类型分析:")
                for col in df.columns:
                    dtype = str(df[col].dtype)
                    non_null_count = df[col].count()
                    null_count = df[col].isnull().sum()
                    unique_count = df[col].nunique()
                    
                    print(f"  {col}:")
                    print(f"    数据类型: {dtype}")
                    print(f"    非空值数量: {non_null_count}")
                    print(f"    空值数量: {null_count}")
                    print(f"    唯一值数量: {unique_count}")
                    
                    # 根据数据类型提供额外信息
                    if pd.api.types.is_numeric_dtype(df[col]):
                        if non_null_count > 0:
                            print(f"    最小值: {df[col].min()}")
                            print(f"    最大值: {df[col].max()}")
                            print(f"    平均值: {df[col].mean():.2f}")
                    elif pd.api.types.is_datetime64_dtype(df[col]):
                        if non_null_count > 0:
                            print(f"    最早日期: {df[col].min()}")
                            print(f"    最晚日期: {df[col].max()}")
                    else:
                        # 对于文本数据，显示前5个唯一值
                        unique_values = df[col].dropna().unique()
                        if len(unique_values) > 0:
                            sample_values = unique_values[:5] if len(unique_values) > 5 else unique_values
                            print(f"    样本值: {list(sample_values)}")
                    print()
                
                # 使用openpyxl检查格式信息（在非只读模式下）
                try:
                    wb_full = openpyxl.load_workbook(file_path, read_only=False)
                    worksheet = wb_full[sheet_name]
                    
                    # 检查合并单元格
                    merged_cells = list(worksheet.merged_cells.ranges)
                    if merged_cells:
                        print(f"合并单元格数量: {len(merged_cells)}")
                        print("合并单元格范围:")
                        for merged_range in merged_cells[:10]:  # 只显示前10个
                            print(f"  {merged_range}")
                        if len(merged_cells) > 10:
                            print(f"  ... 还有 {len(merged_cells) - 10} 个合并单元格")
                    else:
                        print("合并单元格: 无")
                    
                    # 检查公式
                    formula_cells = []
                    for row in worksheet.iter_rows():
                        for cell in row:
                            if cell.value is not None and str(cell.value).startswith('='):
                                formula_cells.append(f"{cell.coordinate}: {cell.value}")
                    
                    if formula_cells:
                        print(f"\n公式单元格数量: {len(formula_cells)}")
                        print("公式示例:")
                        for formula in formula_cells[:5]:  # 只显示前5个
                            print(f"  {formula}")
                        if len(formula_cells) > 5:
                            print(f"  ... 还有 {len(formula_cells) - 5} 个公式")
                    else:
                        print("\n公式: 无")
                    
                    wb_full.close()
                except Exception as format_error:
                    print(f"格式分析出错: {format_error}")
                    print("合并单元格: 无法读取（文件可能受保护或格式复杂）")
                    print("公式: 无法读取（文件可能受保护或格式复杂）")
                
                # 显示数据样本
                print(f"\n数据样本 (前5行):")
                print(df.head().to_string())
                
            except Exception as e:
                print(f"读取工作表 {sheet_name} 时出错: {e}")
            
            print("\n" + "=" * 80 + "\n")
        
        workbook.close()
        
    except Exception as e:
        print(f"分析文件时出错: {e}")

if __name__ == "__main__":
    # 指定要分析的Excel文件路径
    excel_file = "/Users/xuechenglong/Library/CloudStorage/GoogleDrive-alongor0512@gmail.com/我的云端硬盘/WPS+iMac/车险签单清单_2025年3月/保批单业务报表-20250325.xlsx"
    
    analyze_excel_file(excel_file)